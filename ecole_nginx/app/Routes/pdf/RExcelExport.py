"""Routes d'export Excel pour tous les rapports."""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import date
import io, json
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.database import get_db
from app.dependencies.Dependencie import get_current_user
from app.Models.MModels import User

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

router = APIRouter(prefix="/api/v1", tags=["Excel Export"])

# ── Helpers ────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", fgColor="F0F4FA")
BORDER       = Border(
    left=Side(style='thin', color='D0D7E3'),
    right=Side(style='thin', color='D0D7E3'),
    top=Side(style='thin', color='D0D7E3'),
    bottom=Side(style='thin', color='D0D7E3')
)

def make_wb(title: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws

def write_headers(ws, headers: list, row: int = 1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 22

def write_row(ws, values: list, row: int, alt: bool = False):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)

def excel_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ── Schemas ────────────────────────────────────────────────────────────────
class ExcelRegisterReq(BaseModel):
    classe: str
    annee_ac: str
    cycle: str

class ExcelPaymentReq(BaseModel):
    classe: str
    date_debut: str        # annee_academique_id
    versement: str
    date_fin: Optional[str] = None

class ExcelPedagoReq(BaseModel):
    cycle: str
    classe: str
    annee_ac: str
    mois: Optional[str] = "Tous les mois"
    identifiant: Optional[bool] = False

class ExcelGlobalReq(BaseModel):
    type: Optional[str] = "Global"
    date_debut: Optional[date] = None
    date_fin: Optional[date] = None

class ExcelPresenceReq(BaseModel):
    date_debut: str
    date_fin: str
    classe: Optional[str] = "All"

# ── helpers lignes spéciales sans fusion ──────────────────────────────────
def _write_group_row(ws, row, label, fill, font, n_cols, extras=None):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill; c.border = BORDER; c.font = font
    ws.cell(row=row, column=1).value = label
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if extras:
        for col_idx, val in extras:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18

# ── 1. Registre d'inscription — groupé cycle → classe ─────────────────────
@router.post("/export-excel-register")
def export_excel_register(
    req: ExcelRegisterReq,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    from itertools import groupby as igroup

    q = """
        SELECT e.nom, e.prenom, e.identifiant as etudiant_identifiant,
               e.sexe, e.date_de_naissance, e.telephone, e.adresse,
               n.name as niveau_name, a.annee_academique, c.nom_classe
        FROM classes_etudiants ce
        JOIN etudiants e  ON ce.etudiant_id = e.id
        JOIN niveaux n    ON ce.niveau_id = n.id
        JOIN annee_academiques a ON ce.annee_academique_id = a.id
        JOIN classes c    ON ce.classes_id = c.id
        WHERE ce.annee_academique_id = :annee_ac AND ce.status = 1
    """
    params = {"annee_ac": req.annee_ac}
    if req.classe and req.classe not in ("All", "Toutes les classes"):
        q += " AND ce.classes_id = :classe_id"; params["classe_id"] = req.classe
    if req.cycle and req.cycle != "All":
        q += " AND ce.niveau_id = :niveau_id"; params["niveau_id"] = req.cycle
    q += " ORDER BY n.name, c.nom_classe, e.nom"
    rows = [dict(r) for r in db.execute(text(q), params).mappings().all()]

    wb, ws = make_wb("Registre")
    N = 9   # nb colonnes
    CYCLE_F  = PatternFill("solid", fgColor="0B2545")
    CLASSE_F = PatternFill("solid", fgColor="1A4A7A")
    SUBTOT_F = PatternFill("solid", fgColor="E8F4FD")
    GRAND_F  = PatternFill("solid", fgColor="D4A853")
    BF_W     = Font(bold=True, color="FFFFFF", size=10)
    BF_B     = Font(bold=True, color="1A3A5C", size=10)
    BF_G     = Font(bold=True, color="FFFFFF", size=11)

    headers = ["#", "Identifiant", "Nom", "Prénom", "Sexe", "Date naissance", "Téléphone", "Adresse", "Classe"]
    write_headers(ws, headers); ws.freeze_panes = "A2"
    cur = 2; grand_total = 0

    for cycle, c_iter in igroup(rows, key=lambda x: x["niveau_name"]):
        c_list = list(c_iter)
        _write_group_row(ws, cur, f"  ▶  Cycle : {cycle}", CYCLE_F, BF_W, N)
        cur += 1; cycle_tot = 0

        for classe, cl_iter in igroup(c_list, key=lambda x: x["nom_classe"]):
            cl_list = list(cl_iter)
            _write_group_row(ws, cur, f"      Classe : {classe}", CLASSE_F, BF_W, N)
            cur += 1
            for num, r in enumerate(cl_list, 1):
                write_row(ws, [
                    num, r["etudiant_identifiant"], r["nom"], r["prenom"],
                    r["sexe"] or "", str(r["date_de_naissance"] or "")[:10],
                    r["telephone"] or "", r["adresse"] or "", classe
                ], cur, num % 2 == 0); cur += 1
            nb = len(cl_list); cycle_tot += nb
            _write_group_row(ws, cur,
                f"      Sous-total {classe}  ({nb} élève{'s' if nb>1 else ''})",
                SUBTOT_F, BF_B, N, [(N, nb)]); cur += 1

        grand_total += cycle_tot
        _write_group_row(ws, cur, f"  Total Cycle {cycle}  ({cycle_tot} élèves)",
                         CYCLE_F, BF_W, N, [(N, cycle_tot)]); cur += 1

    _write_group_row(ws, cur, f"  ★  TOTAL GÉNÉRAL  ({grand_total} élèves)",
                     GRAND_F, BF_G, N, [(N, grand_total)])
    ws.row_dimensions[cur].height = 22
    auto_width(ws)
    return excel_response(wb, "registre_inscription.xlsx")

# ── 2. Rapport financier ── même logique que le PDF PaymentRepport ──────────
@router.post("/export-excel-paiement")
def export_excel_paiement(
    req: ExcelPaymentReq,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    from app.Models.MFinancials import Paiement as PaiementModel
    from app.Models.MModels import AnneeAcademique
    from app.Routes.pdf.PaymentRepport import (
        extraire_donnees_par_intervalle_not_for_all,
        extraire_donnees_par_intervalle,
        versement_sort_key
    )

    # 1. Récupérer l'année académique (date_debut = id annee_academique)
    annee = db.query(AnneeAcademique).filter(AnneeAcademique.id == req.date_debut).first()
    if not annee:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Année académique introuvable")

    # 2. Récupérer les paiements de cette année
    paiements_query = db.query(PaiementModel.paiement_details).filter(
        PaiementModel.annee_academique == annee.annee_academique
    ).all()
    payment_list = []
    for p in paiements_query:
        try:
            d = json.loads(p.paiement_details) if isinstance(p.paiement_details, str) else p.paiement_details
            if d: payment_list.append(d)
        except Exception: pass

    # 3. Récupérer les élèves (même requête que le PDF)
    query_str = """
        SELECT etudiants.id, etudiants.identifiant, classes.nom_classe,
               niveaux.name, etudiants.nom, etudiants.prenom, etudiants.aide_financiere
        FROM etudiants
        LEFT JOIN classes_etudiants ON classes_etudiants.etudiant_id = etudiants.id
        LEFT JOIN annee_academiques  ON classes_etudiants.annee_academique_id = annee_academiques.id
        LEFT JOIN classes            ON classes_etudiants.classes_id = classes.id
        LEFT JOIN niveaux            ON classes_etudiants.niveau_id = niveaux.id
        WHERE classes_etudiants.annee_academique_id = :annee_id
        AND classes_etudiants.status = 1
    """
    params = {"annee_id": annee.id}
    if req.classe and req.classe.lower() not in ("all", "toutes les classes"):
        query_str += " AND classes.nom_classe = :nom_classe"
        params["nom_classe"] = req.classe
    query_str += " ORDER BY etudiants.nom, niveaux.updated_at"

    data_student_rows = db.execute(text(query_str), params).fetchall()

    class StudentData:
        def __init__(self, row):
            self.id            = row.id
            self.identifiant   = row.identifiant
            self.nom_classe    = row.nom_classe
            self.name          = row.name
            self.nom           = row.nom
            self.prenom        = row.prenom
            self.aide_financiere = row.aide_financiere

    students = [StudentData(r) for r in data_student_rows]

    # 4. Même extraction que le PDF
    versement = req.versement or 'tous les Versements'
    if versement.lower() != 'tous les versements':
        data_payment = extraire_donnees_par_intervalle_not_for_all(
            payment_list, req.date_debut, None, students, req.classe, versement
        )
    else:
        data_payment = extraire_donnees_par_intervalle(
            payment_list, req.date_debut, None, students
        )

    # 5. Construire l'Excel avec regroupement Cycle → Classe → Élèves → Sous-totaux
    all_versement_keys: list = []
    for et in data_payment:
        for k in et.get('entetesVersements', {}).keys():
            if k not in all_versement_keys:
                all_versement_keys.append(k)
    all_versement_keys.sort(key=versement_sort_key)
    versement_headers = all_versement_keys if all_versement_keys else [versement]

    wb, ws = make_wb("Rapport Financier")

    # Styles spéciaux
    CYCLE_FILL   = PatternFill("solid", fgColor="0B2545")
    CLASSE_FILL  = PatternFill("solid", fgColor="1A4A7A")
    SUBTOT_FILL  = PatternFill("solid", fgColor="E8F4FD")
    GRAND_FILL   = PatternFill("solid", fgColor="D4A853")
    CYCLE_FONT   = Font(bold=True, color="FFFFFF", size=10)
    CLASSE_FONT  = Font(bold=True, color="FFFFFF", size=10)
    SUBTOT_FONT  = Font(bold=True, color="1A3A5C", size=10)
    GRAND_FONT   = Font(bold=True, color="FFFFFF", size=11)

    n_cols = 6 + len(versement_headers) + 3  # fixe + versements + total/avance/balance
    tot_col   = 6 + len(versement_headers) + 1   # col index "Total versé" (1-based)
    avnc_col  = tot_col + 1
    bal_col   = tot_col + 2

    # ── En-têtes ──────────────────────────────────────────────────────────────
    fixed   = ["Identifiant", "Nom", "Prénom", "Classe", "Niveau", "Aide financière"]
    headers = fixed + versement_headers + ["Total versé", "Avance", "Balance"]
    write_headers(ws, headers)
    ws.row_dimensions[1].height = 24
    cur_row = 2

    def fmt_val(val, montant_du):
        if val is True:    return f"✓ Payé ({int(montant_du):,})"
        if val is False:   return "✗ Non payé"
        if isinstance(val, (int, float)) and val > 0:
            return f"≈ Avance: {int(val):,}"
        return "—"

    def write_special(ws, row, label, fill, font, subtotals=None):
        """Écrit une ligne spéciale SANS fusion (évite MergedCell read-only)."""
        # Pré-remplir toutes les colonnes de la ligne
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill
            c.border = BORDER
            c.font = font
        # Label dans la première cellule
        c0 = ws.cell(row=row, column=1, value=label)
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # Valeurs numériques dans leurs colonnes respectives
        if subtotals:
            for col_idx, val in subtotals:
                c = ws.cell(row=row, column=col_idx, value=val)
                c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18

    # ── Regrouper ─────────────────────────────────────────────────────────────
    from itertools import groupby as igroup

    data_sorted = sorted(data_payment,
        key=lambda x: (x.get('cycle', ''), x.get('classe_name', ''), x.get('nom', '')))

    grand_total = grand_avance = grand_balance = 0

    for cycle_name, cycle_iter in igroup(data_sorted, key=lambda x: x.get('cycle', 'Non défini')):
        cycle_list = list(cycle_iter)
        cycle_total = cycle_avance = cycle_balance = 0

        # En-tête cycle
        write_special(ws, cur_row, f"  ▶  Cycle : {cycle_name}",
                      CYCLE_FILL, CYCLE_FONT)
        cur_row += 1

        for classe_name, classe_iter in igroup(cycle_list, key=lambda x: x.get('classe_name', 'Non défini')):
            classe_list = list(classe_iter)
            classe_total = classe_avance = classe_balance = 0
            acq_counts = {vk: 0 for vk in versement_headers}

            # En-tête classe
            write_special(ws, cur_row, f"      Classe : {classe_name}",
                          CLASSE_FILL, CLASSE_FONT)
            cur_row += 1

            # Lignes élèves
            for idx, et in enumerate(classe_list):
                pmt    = et.get('paiements', {})
                entetes= et.get('entetesVersements', {})
                t = pmt.get('total', 0) or 0
                a = pmt.get('avance', 0) or 0

                # Balance = 0 si tous les versements sont acquittés (True)
                tous_acquittes = all(pmt.get(vk) is True for vk in versement_headers if vk in pmt)
                b = 0 if tous_acquittes else (pmt.get('balance', 0) or 0)

                classe_total   += t
                classe_avance  += a
                classe_balance += b

                row_vals = [
                    et.get('identifiant', ''), et.get('nom', ''), et.get('prenom', ''),
                    classe_name, cycle_name, et.get('aide_financiere', ''),
                ]
                for vk in versement_headers:
                    val = pmt.get(vk)
                    if val is True: acq_counts[vk] += 1
                    row_vals.append(fmt_val(val, entetes.get(vk, 0)))
                row_vals += [t or '', a or '', b or '']
                write_row(ws, row_vals, cur_row, idx % 2 == 0)
                cur_row += 1

            # Sous-total classe
            nb = len(classe_list)
            subtot_label = f"      Sous-total {classe_name}  ({nb} élève{'s' if nb > 1 else ''})"
            versement_summary = [(6 + j + 1, f"{acq_counts[vk]}/{nb} payés")
                                 for j, vk in enumerate(versement_headers)]
            write_special(ws, cur_row, subtot_label, SUBTOT_FILL, SUBTOT_FONT,
                          versement_summary + [
                              (tot_col,  classe_total),
                              (avnc_col, classe_avance),
                              (bal_col,  classe_balance),
                          ])
            cur_row += 1
            cycle_total   += classe_total
            cycle_avance  += classe_avance
            cycle_balance += classe_balance

        # Sous-total cycle
        nb_c = len(cycle_list)
        write_special(ws, cur_row,
                      f"  Total Cycle {cycle_name}  ({nb_c} élève{'s' if nb_c > 1 else ''})",
                      CYCLE_FILL, CYCLE_FONT,
                      [(tot_col, cycle_total), (avnc_col, cycle_avance), (bal_col, cycle_balance)])
        cur_row += 1
        grand_total   += cycle_total
        grand_avance  += cycle_avance
        grand_balance += cycle_balance

    # ── Total général ──────────────────────────────────────────────────────────
    nb_tot = len(data_payment)
    write_special(ws, cur_row,
                  f"  ★  TOTAL GÉNÉRAL  ({nb_tot} élève{'s' if nb_tot > 1 else ''})",
                  GRAND_FILL, GRAND_FONT, 6 + len(versement_headers),
                  [(tot_col, grand_total), (avnc_col, grand_avance), (bal_col, grand_balance)])
    ws.row_dimensions[cur_row].height = 22

    auto_width(ws)
    # Geler la ligne d'en-têtes
    ws.freeze_panes = "A2"
    return excel_response(wb, f"rapport_financier_{annee.annee_academique}.xlsx")

# ── 3. Rapport pédagogique — même logique que PedagogicRepport.py ──────────
@router.post("/export-excel-pedagogique")
def export_excel_pedagogique(
    req: ExcelPedagoReq,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    from itertools import groupby as igroup
    from app.Models.MModels import AnneeAcademique
    from app.Routes.pdf.PedagogicRepport import (
        get_max_coefficients_for_class, calculer_moyenne_generale
    )

    annee_obj = db.query(AnneeAcademique).filter(AnneeAcademique.id == req.annee_ac).first()
    if not annee_obj:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Année académique introuvable")
    annee_libelle = annee_obj.annee_academique

    # Même SQL que PDF + professeur
    q = """
        SELECT ce.*, c.id as classeId, e.nom, e.prenom, e.identifiant, e.id as etudiant_id,
               c.nom_classe, n.name as niveau_name, coe.data_etudiant,
               (SELECT CONCAT(p.nom, ' ', p.prenom)
                FROM professeurs p
                LEFT JOIN programmes pr ON p.id = pr.professeur_id
                WHERE c.id = pr.class AND ce.niveau_id = pr.niveau_id
                AND pr.annee_academique = :annee_id LIMIT 1) as professeur
        FROM classes_etudiants ce
        JOIN etudiants e         ON ce.etudiant_id = e.id
        JOIN cours_etudiants coe ON coe.etudiant_id = ce.etudiant_id
        JOIN classes c           ON ce.classes_id = c.id
        JOIN niveaux n           ON ce.niveau_id = n.id
        WHERE ce.annee_academique_id = :annee_id
        AND coe.annee_academique = :annee_libelle
        AND ce.status = 1
    """
    params = {"annee_id": req.annee_ac, "annee_libelle": annee_libelle}
    if req.classe and req.classe != "Toutes les classes":
        q += " AND ce.classes_id = :classe_id"; params["classe_id"] = req.classe
    if req.cycle and req.cycle != "All":
        q += " AND ce.niveau_id = :niveau_id"; params["niveau_id"] = req.cycle
    q += " ORDER BY n.name, c.nom_classe, e.nom"
    results = db.execute(text(q), params).mappings().all()

    # Calculer moyennes
    max_coeffs_cache = {}
    students = []
    for row in results:
        if not row["data_etudiant"] or row["data_etudiant"] == "[]": continue
        if "kind" in row["nom_classe"].lower(): continue
        classe_id = row["classeId"]
        if classe_id not in max_coeffs_cache:
            max_coeffs_cache[classe_id] = get_max_coefficients_for_class(
                db, classe_id, annee_libelle, req.mois)
        stats = calculer_moyenne_generale(row["data_etudiant"], row["identifiant"],
                                          max_coeffs_cache[classe_id], req.mois)
        m_passage = 6.5 if "CP" in row["nom_classe"] else 6.0
        students.append({
            "niveau": row["niveau_name"],
            "classe": row["nom_classe"],
            "identifiant": row["identifiant"],
            "nom": row["nom"],
            "prenom": row["prenom"],
            "moyenne": stats[0],
            "m_passage": m_passage,
            "professeur": row["professeur"] or "",
        })

    def mention(m):
        if m >= 9:  return "Très Bien"
        if m >= 8:  return "Bien"
        if m >= 7:  return "Assez Bien"
        if m >= 6:  return "Passable"
        return "Échec"

    # ── Styles ────────────────────────────────────────────────────────────────
    N = 8
    CYCLE_F  = PatternFill("solid", fgColor="0B2545")
    CLASSE_F = PatternFill("solid", fgColor="1A4A7A")
    SUBTIT_F = PatternFill("solid", fgColor="1E5799")   # sous-titres (Premiers / Échecs)
    SUBTOT_F = PatternFill("solid", fgColor="EFF6FF")
    STAT_F   = PatternFill("solid", fgColor="FFF7ED")
    GRAND_F  = PatternFill("solid", fgColor="D4A853")
    BF_W  = Font(bold=True, color="FFFFFF", size=10)
    BF_G  = Font(bold=True, color="FFFFFF", size=11)
    BF_ST = Font(bold=True, color="FFFFFF", size=9)
    GOLD_F  = PatternFill("solid", fgColor="FEF9C3")

    wb, ws = make_wb("Pédagogique")
    mois_label = req.mois if req.mois else "Tous les mois"
    headers = ["Rang", "Identifiant", "Nom", "Prénom",
               f"Moy./10 ({mois_label})", "Mention", "Résultat", "Classe"]
    write_headers(ws, headers)
    ws.freeze_panes = "A2"
    cur = 2; grand_admis = grand_echecs = grand_total = 0

    data_sorted = sorted(students, key=lambda x: (x["niveau"], x["classe"], -x["moyenne"]))

    for cycle, c_iter in igroup(data_sorted, key=lambda x: x["niveau"]):
        c_list = list(c_iter)
        _write_group_row(ws, cur, f"  ▶  Cycle : {cycle}", CYCLE_F, BF_W, N)
        cur += 1; cyc_admis = cyc_echecs = cyc_tot = 0

        for classe, cl_iter in igroup(c_list, key=lambda x: x["classe"]):
            cl_list = list(cl_iter)
            nb = len(cl_list)
            prof = cl_list[0]["professeur"] if cl_list else ""
            is_primary = any(kw in cycle.lower() for kw in ["primaire", "fondamental", "primary"])

            # ── En-tête classe ────────────────────────────────────────────
            classe_info = f"      Classe : {classe}  |  {nb} élève{'s' if nb>1 else ''}"
            if is_primary and prof:
                classe_info += f"  |  Professeur : {prof}"
            _write_group_row(ws, cur, classe_info, CLASSE_F, BF_W, N)
            cur += 1

            # ── Calculer rangs, admis, échecs (sans écrire la liste complète) ──
            cl_moys = []; cl_admis = 0; cl_echecs = 0
            admis_list = []; echecs_list = []

            # cl_list est déjà trié par -moyenne (via data_sorted)
            # On trie explicitement par moyenne desc pour garantir l'ordre
            cl_list_sorted = sorted(cl_list, key=lambda x: -x["moyenne"])

            for rang, et in enumerate(cl_list_sorted, 1):
                moy = et["moyenne"]
                cl_moys.append(moy)
                ok = moy >= et["m_passage"]
                if ok: cl_admis += 1; admis_list.append((rang, et))
                else:  cl_echecs += 1; echecs_list.append((rang, et))

            # ── Section : Premiers (rang 1, ex æquo inclus) ──────────────
            _write_group_row(ws, cur, "      🏆  PREMIERS DE CLASSE", SUBTIT_F, BF_ST, N)
            cur += 1
            if admis_list:
                top_moy = round(admis_list[0][1]["moyenne"], 2)
                premiers = [(r, e) for r, e in admis_list
                            if round(e["moyenne"], 2) == top_moy]
                for rang, et in premiers:
                    write_row(ws, [rang, et["identifiant"], et["nom"], et["prenom"],
                                   et["moyenne"], mention(et["moyenne"]), "Succès", classe], cur, False)
                    for col in range(1, N+1):
                        ws.cell(row=cur, column=col).fill = GOLD_F
                        ws.cell(row=cur, column=col).font = Font(bold=True, color="92400E", size=10)
                    cur += 1
            else:
                c = ws.cell(row=cur, column=1, value="      Aucun admis")
                c.font = Font(italic=True, color="6B7280"); c.fill = SUBTOT_F; cur += 1

            # ── Section : Échecs ──────────────────────────────────────────
            _write_group_row(ws, cur, "      ⚠  ÉLÈVES EN ÉCHEC", SUBTIT_F, BF_ST, N)
            cur += 1
            if echecs_list:
                for rang, et in echecs_list:
                    write_row(ws, [rang, et["identifiant"], et["nom"], et["prenom"],
                                   et["moyenne"], mention(et["moyenne"]), "Échec", classe], cur, False)
                    for col in range(1, N+1):
                        ws.cell(row=cur, column=col).fill = PatternFill("solid", fgColor="FEF2F2")
                        ws.cell(row=cur, column=col).font = Font(color="991B1B", size=10)
                    cur += 1
            else:
                c = ws.cell(row=cur, column=1, value="      Aucun échec")
                c.font = Font(italic=True, color="6B7280"); c.fill = SUBTOT_F; cur += 1

            # ── Statistiques fin de classe ─────────────────────────────────
            moy_cl   = round(sum(cl_moys)/len(cl_moys), 2) if cl_moys else 0
            pct_reuss = round(cl_admis / nb * 100, 1) if nb else 0
            pct_echec = round(cl_echecs / nb * 100, 1) if nb else 0
            stat_label = (
                f"      📊  Taux réussite : {cl_admis}/{nb} ({pct_reuss}%)   |   "
                f"Taux échec : {cl_echecs}/{nb} ({pct_echec}%)   |   "
                f"Moy. classe : {moy_cl}/10"
            )
            _write_group_row(ws, cur, stat_label, STAT_F,
                             Font(bold=True, color="7C3AED", size=10), N,
                             [(5, moy_cl)])
            ws.row_dimensions[cur].height = 20
            cur += 1

            # Ligne vide entre classes
            cur += 1
            cyc_admis += cl_admis; cyc_echecs += cl_echecs; cyc_tot += nb

        # Total cycle
        pct_c = round(cyc_admis/cyc_tot*100, 1) if cyc_tot else 0
        _write_group_row(ws, cur,
            f"  Total Cycle {cycle} — {cyc_tot} élèves | Admis : {cyc_admis} ({pct_c}%) | Échecs : {cyc_echecs} ({100-pct_c}%)",
            CYCLE_F, BF_W, N)
        cur += 1; cur += 1   # ligne vide entre cycles
        grand_admis += cyc_admis; grand_echecs += cyc_echecs; grand_total += cyc_tot

    # Total général
    pct_g = round(grand_admis/grand_total*100, 1) if grand_total else 0
    _write_group_row(ws, cur,
        f"  ★  TOTAL GÉNÉRAL — {grand_total} élèves | Admis : {grand_admis} ({pct_g}%) | Échecs : {grand_echecs} ({100-pct_g}%)",
        GRAND_F, BF_G, N)
    ws.row_dimensions[cur].height = 22
    auto_width(ws)
    return excel_response(wb, f"rapport_pedagogique_{annee_libelle}.xlsx")

# ── 4. Rapport global ──────────────────────────────────────────────────────
# Même logique que le PDF : les versements sont dans le JSON info_paiement,
# on filtre par les clés de date du JSON (format "dd-mm-YYYY HH:MM")
@router.post("/export-excel-global")
def export_excel_global(
    req: ExcelGlobalReq,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    from app.Models.MFinancials import Paiement as PaiementModel
    from datetime import datetime as dt, time as dtime
    from sqlalchemy import and_

    # Même filtre SQL que le PDF
    date_debut = dt.combine(req.date_debut, dtime.min) if req.date_debut else dt(2000, 1, 1)
    date_fin   = dt.combine(req.date_fin,   dtime.max) if req.date_fin   else dt(2099, 12, 31)

    paiements_query = db.query(PaiementModel.paiement_details).filter(
        and_(
            PaiementModel.created_at <= date_fin,
            PaiementModel.updated_at >= date_debut
        )
    ).all()

    # Parser les JSON
    payment_list = []
    for p in paiements_query:
        try:
            data = json.loads(p.paiement_details) if isinstance(p.paiement_details, str) else p.paiement_details
            if data:
                payment_list.append(data)
        except Exception:
            pass

    # Extraire les versements dans l'intervalle (même logique que le PDF)
    lignes = []
    for value in payment_list:
        pd_inner = value.get('paiement_details', {})
        data_student = pd_inner.get('details_etudiant', {})
        info_paiement = pd_inner.get('info_paiement', {})

        for date_str, donnees in info_paiement.items():
            if donnees.get('status') == 'retourné':
                continue
            try:
                date_cle = dt.strptime(date_str, '%d-%m-%Y %H:%M')
            except ValueError:
                try:
                    date_cle = dt.strptime(date_str, '%d-%m-%Y %H:%M:%S')
                except ValueError:
                    continue

            if date_debut <= date_cle <= date_fin:
                lignes.append({
                    'date_paiement': date_cle.strftime('%d/%m/%Y %H:%M'),
                    'nom':           data_student.get('nom', ''),
                    'prenom':        data_student.get('prenom', ''),
                    'identifiant':   data_student.get('identifiant', ''),
                    'classe':        data_student.get('classe', ''),
                    'niveau':        data_student.get('niveau', ''),
                    'aide':          donnees.get('aide_financiere', ''),
                    'depot':         donnees.get('depot', 0) or 0,
                    'total_verse':   donnees.get('total_verse', 0) or 0,
                    'total_annuel':  donnees.get('total_annuel', 0) or 0,
                    'balance':       donnees.get('balance', 0) or 0,
                    '_raw':          donnees,   # pour parse_situation
                })

    wb, ws = make_wb("Rapport Global")
    headers = [
        "Date versement", "Nom", "Prénom", "Identifiant",
        "Classe", "Niveau", "Aide financière",
        "Dépôt", "Total versé", "Total annuel", "Balance",
        "Versements acquittés", "Avance sur", "Situation"
    ]
    write_headers(ws, headers)

    def parse_situation(donnees):
        sp = donnees.get('status_paiement')
        status = donnees.get('status')
        balance = donnees.get('balance') or 0
        total_verse = donnees.get('total_verse') or 0
        total_annuel = donnees.get('total_annuel') or 0

        acquittes, avances = [], []
        if isinstance(sp, list):
            for s in sp:
                if str(s).startswith('Acqt:'):
                    acquittes.append(str(s).replace('Acqt: ', '').strip())
                elif str(s).startswith('Avns:'):
                    avances.append(str(s).replace('Avns: ', '').strip())

        # Situation lisible
        if status == 'Acquitté' or (total_annuel and total_verse >= total_annuel):
            situation = 'Acquitté'
        elif acquittes and avances:
            situation = f'Partiellement acquitté + avance'
        elif acquittes:
            situation = 'Versement(s) acquitté(s)'
        elif avances:
            situation = 'Avance partielle'
        elif balance and balance > 0:
            situation = 'En cours'
        elif status and status not in (0, None, ''):
            situation = str(status)
        else:
            situation = 'En cours'

        return ', '.join(acquittes), ', '.join(avances), situation

    for i, r in enumerate(sorted(lignes, key=lambda x: x['date_paiement'], reverse=True), 2):
        acqt, avns, sit = parse_situation(r['_raw'])
        write_row(ws, [
            r['date_paiement'], r['nom'], r['prenom'], r['identifiant'],
            r['classe'], r['niveau'], r['aide'],
            r['depot'], r['total_verse'], r['total_annuel'], r['balance'],
            acqt, avns, sit
        ], i, i % 2 == 0)
    auto_width(ws)
    fname = f"rapport_global_{req.date_debut or 'all'}_{req.date_fin or 'all'}.xlsx"
    return excel_response(wb, fname)

# ── 5. Rapport présences ───────────────────────────────────────────────────
@router.post("/export-excel-presence")
def export_excel_presence(
    req: ExcelPresenceReq,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    sql = text("""
        SELECT e.nom, e.prenom, e.identifiant,
               c.nom_classe,
               COUNT(CASE WHEN pr.valeur = 1 THEN 1 END) as nb_present,
               COUNT(CASE WHEN pr.valeur = 0 THEN 1 END) as nb_absent,
               COUNT(*) as total_jours
        FROM presences pr
        JOIN etudiants e ON pr.etudiant_id = e.id
        LEFT JOIN classes c ON pr.classes_id = c.id
        WHERE pr.date_daujourdhui BETWEEN :d1 AND :d2
        GROUP BY e.id, e.nom, e.prenom, e.identifiant, c.nom_classe
        ORDER BY c.nom_classe, e.nom
    """)
    params = {"d1": req.date_debut, "d2": req.date_fin}
    rows = db.execute(sql, params).mappings().all()

    wb, ws = make_wb("Présences")
    headers = ["Nom", "Prénom", "Identifiant", "Classe", "Présents", "Absents", "Total jours", "Taux présence"]
    write_headers(ws, headers)
    for i, r in enumerate(rows, 2):
        total = r["total_jours"] or 1
        taux = f"{round(r['nb_present'] / total * 100)}%" if total else "–"
        write_row(ws, [
            r["nom"], r["prenom"], r["identifiant"], r["nom_classe"] or "",
            r["nb_present"], r["nb_absent"], r["total_jours"], taux
        ], i, i % 2 == 0)
    auto_width(ws)
    return excel_response(wb, f"rapport_presence_{req.date_debut}_{req.date_fin}.xlsx")
